import openpyxl
from openpyxl import *
from pandas import *
import pandas


print(openpyxl.__version__)
fname = "MOVC_CA_GENNAIO_2016.xlsx"
output = "sorted.xlsx"
wb = load_workbook(fname)
names = wb.get_sheet_names()
print(names)
wb._sheets.sort(key = lambda ws: ws.title)
names_after = wb.get_sheet_names()
print(names_after)
wb.save(output)

# fname = "sample_new.xlsx"
# output = "output_new.xlsx"
# wb = load_workbook(fname)
# ws = wb.active
# ws2 = wb.copy_worksheet(ws)
# ws2.title = 'Sassari'
#
# ws2['D5'] = 3
#
# row_indexes = [5,6]
# col_indexes = [4,6,8]
# values = [10,11,12]
#
# for r in row_indexes:
#     counter = 0
#     for c in col_indexes:
#         ws2.cell(row=r,column=c, value=values[counter])
#         counter += 1
# wb.save(output)

# data = ws.values
# cols = next(data)[1:]
# data = list(data)
# idx = [r[0] for r in data]
# data = (pandas.islice(r, 1, None) for r in data)
# df = DataFrame(data, index=idx, columns=cols)




# import xlrd
#
#
#
#
# fname = "sample.xls"
# bk = xlrd.open_workbook(fname)
# shxrange = range(bk.nsheets)
# sh = bk.sheet_by_index(0)
# # try:
# #     sh = bk.sheet_by_name("Sheet1")
# # except:
# #     print("no sheet in %s named Sheet1" % fname)
#
# nrows = sh.nrows
# ncols = sh.ncols
# print ("nrows %d, ncols %d" % (nrows,ncols))
#
# cell_value = sh.cell_value(0,1)
# print(cell_value)
#
# # row_list = []
# # for i in range(1,nrows):
# #     row_data = sh.row_values(i)
# #     row_list.append(row_data)