__author__ = 'pg'
from openpyxl import *
from movc.movc_parser import MovcParser
from movc.movc_fields import *
import locale
import calendar

class MovXL:


    def __init__(self, template_file_path, mapper_file, movc_file):
         self.template = load_workbook(template_file_path)
         self.parser = MovcParser(movc_file, mapper_file)
         self.parser.open_session()

    def get_movc(self):
        return(self.template)

    def get_parser(self):
        return(self.parser)

    def build_xl(self, filepath):

        blocks = self.parser.get_movc_blocks()
        active_ws = self.template.active

        idx = 1
        for b in blocks:
            consinstency = []
            movements = []
            # print(self.parser.get_comune(b)[0])

            ###info to write ###
            provincia = self.parser.get_province(b)
            comune, codice_comune = self.parser.get_comune(b)
            codice_comune = '0' + codice_comune
            year = self.parser.get_year(b)
            month = self.parser.get_month(b)

            info_comune = {'provincia': provincia, 'comune':comune, 'codice_comune': codice_comune, 'year': year,
                           'month': month}

            ##workbook generation ###
            ws = active_ws
            ws.title = comune.upper()

            for code in CAPACITY_CODES:
                consinstency.append(self.parser.get_capacity(code, b))

            for origin in ORIGIN_CODES:
                mov  = self.parser.get_movements(origin, b)
                if (isinstance(mov, list)):
                    movements.append(mov)

            active_ws = self.populate_sheet(ws, info_comune, consinstency, movements)
            idx += 1
            if idx <= len(blocks):
                new_ws = self.template.copy_worksheet(ws)
                active_ws = new_ws

        self.template._sheets.sort(key=lambda ws: ws.title)
        self.template.save(filepath)


    def populate_sheet(self, sheet, info_comune, consistency, movements):
        sheet['A1'] = (info_comune['provincia']).upper()
        sheet['A2'] = (info_comune["comune"]).upper()
        sheet['B2'] = info_comune["year"]
        sheet['B3'] = None
        sheet['D2'] = info_comune["month"]
        sheet['D3'] = info_comune["codice_comune"]

        locale.setlocale(locale.LC_ALL, 'ita')
        month_name = (calendar.month_name[int(info_comune["month"])]).upper()
        sheet['A3'] = month_name

        ###flattening lambda function for movements
        flatten = lambda l: [item for sublist in l for item in sublist]

        rcounter = 0
        for r in SHEET_ROW_IDX:
            if r == 16:
                rcounter = 0
            current_collection = consistency[rcounter] if r < CONS_UPPER_ROW_LIMIT else movements[rcounter]
            column_idx = CONS_COL_IDX if r < CONS_UPPER_ROW_LIMIT else MOV_COL_IDX
            if r >= CONS_UPPER_ROW_LIMIT:
                current_collection = flatten(current_collection)
            ccounter = 0
            for c in column_idx:
                sheet.cell(row=r, column=c, value=current_collection[ccounter])
                ccounter += 1
            rcounter += 1
        return sheet

    def add_summary(self, filepath):
        wb =  load_workbook(filepath)
        sheets = wb.get_sheet_names()
        last = wb.get_sheet_by_name(sheets[-1])


        summary = wb.copy_worksheet(last)
        ###Info di riepilogo####
        summary.title = "Riepilogo"
        summary['A2'] = "Riepilogo"
        summary['B3'] = None
        summary['C3'] = None
        summary['D3'] = None
        for r in SHEET_ROW_IDX:
            column_idx = CONS_COL_IDX if r < CONS_UPPER_ROW_LIMIT else MOV_COL_IDX
            data_list = []
            summary_data = []
            for sheet_name in sheets:
                current_sheet = wb.get_sheet_by_name(sheet_name)
                data = [current_sheet.cell(row=r, column = i).value for i in column_idx]
                data_list.append(data)
            #here we sum the same row of each sheet
            summary_data = [sum(x) for x in zip(*data_list)]
            counter = 0
            for k in column_idx:
                summary.cell(row=r, column=k, value=summary_data[counter])
                counter += 1

        wb.save(filepath)

        return(data)


