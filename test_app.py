# import os
# import glob
# import sys
# from movc import *
# import csv
# import datetime
# from movc.movutility import MovUtility
# from movc.new_movc import Movc
# from movc.province_data import *
# import movc.consts


from openpyxl import *

#rom openpyxl import *

import os
import copy
print(os.getcwd())

filepath = os.path.join("config", "movc_base_2012.xlsx")
output = os.path.join("output", "pippo.xlsx")
wb = load_workbook(filepath)

ws = wb.active



ws1 = wb.create_sheet("Armungia")
ws1.append(ws['A1:A10'])
# ws1['A1':'A98']


wb.save(output)

#openpyxl.worksheet.copier.Worksheet(ws, ws1)

# d = {'one': 1, 'two':2}
#
# print(d)
# print('two' in list(d.keys()))
#
# w = {x: x**2 for x in (3,6, 9)}
# print(w)
#
# print(os.getcwd())
# f = open("config/movc_project.txt")
# print(f.readline())
#
# os.chdir("/Users/piepalla")
# print(os.getcwd())
#
# print(os.listdir())
# print((glob.glob('.a*')))
#
# for k in d.keys():
#     print('key = ', k, ' value = ', d[k])
#
# x1 = 0
# def foo(x  = x1):
#     print(x)
#
# print(foo(10))
# print(foo())
# x1 = 33
# print("x1 = ", x1)
# print(foo())
#
# l = [x**2 for x in [5,6,7]]
# print(l)
#
# a = set('abracadbra')
# print(a)
#
# def foo(x=0, y=0, z=0):
#     print("x, y, z = ", x, y, z)
#
# foo(**{'x': 10,'y': 30})
#
# a = foo
# print(a)
# b = foo()
# print(b)
#
# print(sys.path)

#####
# with open('config/mapping.csv', newline = '') as csvfile:
#     mapreader = csv.reader(csvfile, delimiter = ';')
#     for row in mapreader:
#         print(', '.join(row))

# path = "C:\\Users\\piepalla\\PycharmProjects\\movc_project\\movs"
# mu = MovUtility()
# mu.merge_movs(path, "merged_files", "")



#print(mapper.ix[1:3, :])
#mapper.columns = ["prov_new", "com_new", "pro_old", "com_ old"]



# print(mapper.shape)
# print(mapper.__class__)
# print("Iterating over a dataframe")
# for r in mapper.iterrows():
#     print(r)
#print(mapper.shape)
# print(mapper.head(13))
# map_path = "C:\\Users\\piepalla\\PycharmProjects\\new_movc\\config\\mapping.csv"
# mov_path = "C:\\Users\\piepalla\\PycharmProjects\\new_movc\\output\\movc_nu_82016.txt"
# mymov = Movc(8, 2016)
# mymov.set_mapper(map_path)
# mapper = mymov.get_mapper()


#
# print(mapper.ix[2, 0:2])
# IDX = (1, 2, 5, 4)
# print(mapper.ix[0:5, IDX])


### Movc creation ###
# input_path = "C:\\Users\\piepalla\\PycharmProjects\\new_movc\\merged_files\\all_082016.txt"
# output_path = "C:\\Users\\piepalla\\PycharmProjects\\new_movc\\output"
# mymov.create_movc("sud sardegna", input_path, output_path)
#mymov.validate(mov_path)
######

# source_path = "C:\\Users\\piepalla\\PycharmProjects\\movc_project\\movs\movc_ca_082016.txt"
# f = open(source_path, 'r')
# for row in f:
#     r = f.readline()
#     print("row length: " + str(len(r)))
#print(mm)

# year = 2016
# month = 1
#
# input, output = MovUtility.config_io_paths(year, month)
# print(input)
# print(output)
#
# print(movc.consts.BASEDIR)
# print(movc.consts.MAPPING_FILE)