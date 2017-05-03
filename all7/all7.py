__author__ = 'pg'

from openpyxl import *
from movc.province_data import *
from all7.all7_consts import *

class All7:
    def __init__(self, template_file_path, year, province):
        self.year = year
        self.template = load_workbook(template_file_path)
        self.province = province


    def load_xl(self, movc_xl_path):
        self.movc = load_workbook(movc_xl_path, data_only=True)
        self.sheets = self.movc.get_sheet_names()
        if self.sheets[-1] == 'Riepilogo':
            self.sheets = self.sheets[0:-1]


    def get_alberghi(self, type='arrivi', category='residenti'):
        tot = 0
        for sheet_name in self.sheets:
            current_sheet = self.movc.get_sheet_by_name(sheet_name)
            # print(current_sheet.title)
            RANGE = RESIDENTS_RANGE if category == 'residenti' else NON_RESIDENTS_RANGE
            for i in RANGE:
                idx = 'O' + str(i) if type == 'arrivi' else 'P' + str(i)
                # print(current_sheet[idx].value)
                tot += int(current_sheet[idx].value)
        print(tot)
        return(tot)

    def get_alloggi(self, type='arrivi', category='residenti'):
        tot = 0
        #sheet_name = self.sheets[0]
        for sheet_name in self.sheets:
            current_sheet = self.movc.get_sheet_by_name(sheet_name)
            # print(current_sheet.title)
            RANGE = RESIDENTS_RANGE if category == 'residenti' else NON_RESIDENTS_RANGE
            for row_idx in RANGE:
                for k in ALLOGGI_DICT.keys():
                    value_idx = 0 if (type == 'arrivi') else 1
                    idx = ALLOGGI_DICT[k][value_idx] + str(row_idx)
                    tot += current_sheet[idx].value
                    #print("row id: " + str(row_idx) + " " + str(tot_arrivi))
        return(tot)

    def get_campeggi(self, type="arrivi", category='residenti'):
        tot = 0
        for sheet_name in self.sheets:
            current_sheet = self.movc.get_sheet_by_name(sheet_name)
            RANGE = RESIDENTS_RANGE if category == 'residenti' else NON_RESIDENTS_RANGE
            for row_idx in RANGE:
                for k in CAMPEGGI_DICT.keys():
                    value_idx = 0 if (type == "arrivi") else 1
                    idx = CAMPEGGI_DICT[k][value_idx] + str(row_idx)
                    # print(idx)
                    tot += current_sheet[idx].value
        return(tot)

    def get_altri_alloggi(self, type='arrivi', category='residenti'):
        tot = 0
        for sheet_name in self.sheets:
            current_sheet = self.movc.get_sheet_by_name(sheet_name)
            RANGE = RESIDENTS_RANGE if category == 'residenti' else NON_RESIDENTS_RANGE
            for row_idx in RANGE:
                for k in ALTRI_ALLOGGI_DICT.keys():
                    value_idx = 0 if (type == "arrivi") else 1
                    idx = ALTRI_ALLOGGI_DICT[k][value_idx] + str(row_idx)
                    # print(idx)
                    tot += current_sheet[idx].value
        return(tot)


    def get_giornate_letto(self):
        tot = 0
        for sheet_name in self.sheets:
            # print(sheet_name)
            current_sheet = self.movc.get_sheet_by_name(sheet_name)
            idx = 'P11'
            tot += int(current_sheet[idx].value)
        return(tot)


    def get_giornate_camere(self, type = 'disponibili'):
        tot = 0
        idx = 'P12' if (type == 'disponibili') else 'P13'
        for sheet_name in self.sheets[0:3]:
            # print(sheet_name)
            current_sheet = self.movc.get_sheet_by_name(sheet_name)
            tot += int(current_sheet[idx].value)
        return(tot)


    def build_xl(self, movc_dir, output_file):

        template_sheets = self.template.get_sheet_names()
        file_names = []
        buffer_residents = []
        buffer_no_residents = []
        totali_res = [0]*8
        totali_no_res = [0]*8
        totali = [0]*8
        totali_giornate = [0]*3

        try:
            file_names = os.listdir(movc_dir)
        except NotADirectoryError():
            print("Wrong directory! Please enter the correct movc directory name")
            return
        self.check_movc_files(movc_dir)


        for file in file_names:
            file_path = os.path.join(movc_dir, file)
            self.load_xl(file_path)
            if (self.check_province(self.movc)!= self.province and self.check_year(self.movc) != self.year):
                print("Wrong province or year selected!")
            month = self.movc.active['A3'].value
            #### Residents
            idx_res = ALL7_RESIDENTS_RANGE[MONTHS_DICT[month]]
            arrivi_alberghi_res = self.get_alberghi()

            presenze_alberghi_res = self.get_alberghi(type='presenze')
            arrivi_alloggi_res = self.get_alloggi()
            presenze_alloggi_res = self.get_alloggi(type='presenze')
            arrivi_campeggi_res = self.get_campeggi()
            presenze_campeggi_res = self.get_campeggi(type = 'presenze')
            arrivi_altri_alloggi_res = self.get_altri_alloggi()
            presenze_altri_alloggi_res = self.get_altri_alloggi(type='presenze')

            buffer_residents = [arrivi_alberghi_res, presenze_alberghi_res, arrivi_alloggi_res, presenze_alloggi_res, arrivi_campeggi_res, presenze_campeggi_res,
                                arrivi_altri_alloggi_res, presenze_altri_alloggi_res]
            totali_res = [sum(x) for x in zip(totali_res, buffer_residents)]

            ###NON Residents
            idx_no_res = ALL7_NON_RESIDENTS_RANGE[MONTHS_DICT[month]]
            arrivi_alberghi_no_res = self.get_alberghi(category='non residenti')
            presenze_alberghi_no_res = self.get_alberghi(type='presenze', category='non residenti')
            arrivi_alloggi_no_res = self.get_alloggi(category='non residenti')
            presenze_alloggi_no_res = self.get_alloggi(type='presenze', category='non residenti')
            arrivi_campeggi_no_res = self.get_campeggi(category='non residenti')
            presenze_campeggi_no_res = self.get_campeggi(type='presenze', category='non residenti')
            arrivi_altri_alloggi_no_res = self.get_altri_alloggi(category='non residenti')
            presenze_altri_alloggi_no_res = self.get_altri_alloggi(type='presenze', category='non residenti')

            buffer_no_residents = [arrivi_alberghi_no_res, presenze_alberghi_no_res, arrivi_alloggi_no_res, presenze_alloggi_no_res, arrivi_campeggi_no_res, presenze_campeggi_no_res,
                                   arrivi_altri_alloggi_no_res, presenze_altri_alloggi_no_res]

            totali_no_res = [sum(x) for x in zip(totali_no_res, buffer_no_residents)]
            totali = [sum(z) for z in zip(totali_res, totali_no_res)]

            self.template.active = 0

            buff_idx = 0
            for c in ALL7_COL1:
                self.template.active.cell(row=idx_res, column=c, value=buffer_residents[buff_idx])
                self.template.active.cell(row=idx_no_res, column=c, value=buffer_no_residents[buff_idx])
                buff_idx += 1

            ### Totals for residents and non residents
            tot_idx = 0
            for c in ALL7_COL1:
                self.template.active.cell(row = TOT_RES_IDX, column=c, value=totali_res[tot_idx])
                self.template.active.cell(row=TOT_NO_RES_IDX, column=c, value=totali_no_res[tot_idx])
                self.template.active.cell(row=TOT_IDX, column=c, value=totali[tot_idx])
                tot_idx += 1

            ### Available bed days
            self.template.active = 1
            giornate_letto = self.get_giornate_letto()
            giornate_camere = self.get_giornate_camere()
            giornate_camere_occupate = self.get_giornate_camere(type='occupate')
            buffer_giornate = [giornate_letto, giornate_camere, giornate_camere_occupate]

            totali_giornate = [sum(x) for x in zip(totali_giornate, buffer_giornate)]

            day_idx = ALL7_DAYS_RANGE[MONTHS_DICT[month]]
            g_idx = 0
            for c in ALL7_COL2:
                self.template.active.cell(row = day_idx, column = c, value = buffer_giornate[g_idx])
                g_idx += 1

            tot_day_idx = 0
            for c in ALL7_COL2:
                self.template.active.cell(row=TOT_DAYS_IDX, column = c, value = totali_giornate[tot_day_idx])
                tot_day_idx += 1

        self.template.save(output_file)


    def check_movc_files(self, movc_dir):
        if (not os.path.isdir(movc_dir)):
            print("invalid movc directory!")
            return
        files = os.listdir(movc_dir)
        if len(files) != NUM_OF_MOVC:
            print("invalid number of movc modules")
            return


    def check_province(self, workbook):
        ws = workbook.active
        province = ws[PROVINCE_CELL].value
        return province

    def check_year(self, workbook):
        ws = workbook.active
        year = ws[YEAR_CELL].value
        return year