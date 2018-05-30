__author__ = 'pg'
import unittest, logging
from ctt4.ctt4 import *
from movc.province_data import *
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
import os

class TestCtt4(unittest.TestCase):
    def setUp(self):
        ### Set Logger ###
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        ch = logging.StreamHandler()
        # console handler with level debug
        ch.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        ch.setFormatter(formatter)
        self.logger.addHandler(ch)

        self.template_path = r"C:\Users\piepalla\PycharmProjects\new_movc\config\ctt4_base.xlsx"
        self.mapping_path = r"C:\Users\piepalla\PycharmProjects\new_movc\config\mapping.csv"
        self.file_example = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\input\2017\CA_CTT4_2017.xlsx"
        self.input_dir = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\input\\"
        self.ctt4 = Ctt4(self.template_path, self.mapping_path, 2017, "PROVINCIA DI SASSARI")
        #self.all7.load_xl(self.file_example)

    def tearDown(self):
        pass

    def test_init(self):
        # self.assertIsNotNone(self.ctt4)
        # self.assertEqual(self.ctt4.year, 2017)
        # self.assertEqual(self.ctt4.province, "PROVINCIA DI SASSARI")
        self.assertIsNotNone(self.ctt4.mapping)
        self.logger.debug(self.ctt4.mapping.head(3))
        sub = self.ctt4.mapping[self.ctt4.mapping.b == 292]

        self.logger.info(sub.d)
        # self.logger.info("column class")
        # self.logger.info(type(sub.d))
        # self.logger.info("for cycle")

        self.logger.debug(self.ctt4.mapping.iloc[self.ctt4.mapping.b == 292, 0:4])

    def test_get_municipalities(self):
        province_municipalities = self.ctt4.get_municipalities(292)
        self.logger.info(province_municipalities)
        self.assertEqual(len(province_municipalities), 17) # città metropolitana
        province_municipalities = self.ctt4.get_municipalities(111) # sud sardegna
        self.assertEqual(len(province_municipalities), 107) #
        province_municipalities = self.ctt4.get_municipalities(95) # oristano
        self.assertEqual(len(province_municipalities), 87)
        province_municipalities = self.ctt4.get_municipalities(91) # nuoro
        self.assertEqual(len(province_municipalities), 74)
        province_municipalities = self.ctt4.get_municipalities(90) # sassari
        self.assertEqual(len(province_municipalities), 92)




        # for k,v in NEW_CODES.items():
        #     self.logger.info(k)
        #     self.logger.info(self.ctt4.get_municipalities(v))

    def test_get_municipality_code(self):
        filename = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\output\ss_ctt4_2017.xlsx"
        wb = load_workbook(filename)
        ws = wb.active
        municipality = ws['B10'].value

        code = self.ctt4.get_municipality_code(municipality)
        self.logger.info(code)
        self.logger.info(type(code))


    def test_mapping(self):
        m = self.ctt4.mapping
        self.logger.info(m.head(3))
        m_column = m[m.d == "Arzachena"].d
        self.logger.info(m_column)
        self.logger.info(type(m_column))
        m_value = m_column.iloc[0]
        self.logger.info(m_value)
        self.logger.info(type(m_value))
        #self.assertEqual("006", code)

        # municipality = ws['B70'].value
        # code = self.ctt4.get_municipality_code(municipality)
        # self.logger.info(code)
        # self.assertEqual(code, "051")
        # municipality_map = self.ctt4.mapping[self.ctt4.mapping.d == "Arzachena"]
        # self.logger.info(municipality_xls)
        # self.logger.info(municipality_map)
        # municipality2 = self.ctt4.mapping[self.ctt4.mapping.d == municipality_xls].c
        # self.assertEqual(int(municipality2), 90006)
        # municipality_xls = ws['B9'].value
        # self.logger.info(municipality_xls)
        # municipality2 = self.ctt4.mapping[self.ctt4.mapping.d == municipality_xls].d
        # self.logger.info(municipality2)
        #self.assertEqual(municipality_xls, self.ctt4.mapping[self.ctt4.mapping.d == municipality_xls].d)




    def test_build_ctt4(self):
        province_name = 'sassari'
        year = 2017
        input_dir = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\input\\"
        output_dir = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\output\\"
        self.ctt4.build_ctt4(province_name, year, input_dir, output_dir)

        # wb = load_workbook(filename=self.file_example)
        # sheets = wb.get_sheet_names()
        # self.logger.info(sheets)
        # i = 0
        # for s in sheets[0:2]:
        #     i = i +1
        #     active_sheet = wb.get_sheet_by_name(s)
        #     print("sez " + str(i))
        #     #print(active_sheet.cell(row = 10, column = 2).value)
        #     for c in active_sheet['B'][8:]:
        #         if c.value == "TOTALE":
        #             break
        #         print(str(c.row) + " " + str(c.value))


    def test_get_matching_rows(self):
        province_municipalities = self.ctt4.get_municipalities(292) # città metropolitana
        row_list1 = self.ctt4.get_matching_rows(province_municipalities,  self.file_example)[CTT4_SHEET1]
        # # row_list2 = self.ctt4.get_matching_rows(province_municipalities,  self.file_example)[CTT4_SHEET2]
        # # self.assertIsNotNone(row_list1)
        # self.assertIsNotNone(row_list2)

        row1 = row_list1[0]
        columns = [x.column for x in row1]
        self.logger.info(columns)
        # self.assertEqual(len(row_list1), 17)
        # self.assertEqual(len(row_list2), 17)
        # self.logger.info(len(row_list1))
        # self.logger.info(len(row_list2))
        # self.logger.info(row_list1)
        # self.logger.info(row_list2)
        # single_cell1 = row_list1[0][1]
        # single_cell2 = row_list2[0][1]
        # self.logger.info(single_cell1.row)
        # self.logger.info(single_cell1.value)
        # self.logger.info(single_cell1.coordinate)
        # self.logger.info(single_cell2.row)
        # self.logger.info(single_cell2.value)
        # self.logger.info(single_cell2.coordinate)


    def test_writeCTT4(self):
        # province_name = "cagliari"
        # year = 2017
        # output_dir = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\output\\"
        # res = self.ctt4.writeCTT4(province_name, year, output_dir)
        # self.assertIsNotNone(res)
        expected_filepath = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\output\\ca_ctt4_2017bb.xlsx"
        # self.assertEqual(res, expected_filepath)
        file1 = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\output\ss_ctt4_2017.xlsx"
        file2 = r"C:\Users\piepalla\PycharmProjects\new_movc\img\istat.png"
        wb = load_workbook(file1)
        ws = wb.active
        img = Image(file2)
        ws.add_image(img, 'A1')
        wb.save(file1)


    def test_compute_totals(self):
        file1 = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\output\2017\ca_ctt4_2017.xlsx"
        #wb = load_workbook(file1)
        #ws = wb.get_sheet_by_name('Sez 1')
        self.logger.info(self.ctt4.compute_totals(file1))



    def test_styles(self):
        file_path = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\input\2017\style_input.xlsx"
        wb = load_workbook(file_path)
        thin_border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"),
                             top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
        ws = wb.active
        ws["B3"].border = thin_border
        ws["A5"].border = thin_border
        output = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\output\2017\style_test.xlsx"

        wb.save(output)


    def test_apply_formula(self):
        cols = ["AB", "AC", "AD", "AE"]
        self.logger.info("***** Sez 1 ****")
        for c in cols:
            f = self.ctt4.apply_formula('Sez 1', 9, c)
            self.logger.info(f)

        cols = ["P", "Q", "R", "S"]
        self.logger.info("***** Sez 2 ****")
        for c in cols:
            f = self.ctt4.apply_formula('Sez 2', 9, c)
            self.logger.info(f)



    def test_sum_formula(self):
        file1 = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\examples\formula1.xlsx"
        output = r"C:\Users\piepalla\PycharmProjects\new_movc\ctt4\examples\out1.xlsx"
        wb = load_workbook(file1)
        ws = wb.active
        self.logger.info(ws["A4"].value)
        self.logger.info(ws["B4"].value)
        self.logger.info(ws["D4"].value)

        formula = "=SUM(A4,B4,D4)"
        ws["H4"] = formula
        wb.save(output)
####### remove useless sheets  - to do ####
        # sheets = self.ctt4.template.get_sheet_names()
        # self.logger.debug(sheets)
        # self.logger.debug(sheets[2:7])
        #
        # for s in sheets[2:7]:
        #      std = self.ctt4.template.get_sheet_by_name(s)
        #      self.ctt4.template.remove_sheet(std)
        #
        # self.logger.info("sheets after cleaning operation")
        # self.logger.debug(self.ctt4.template.get_sheet_names())
        # self.ctt4.template.save(r"C:\Users\piepalla\PycharmProjects\new_movc\config\ctt4_base.xlsx")
###########################################

