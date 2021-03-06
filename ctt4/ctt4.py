__author__ = 'pg'

from openpyxl import *
from openpyxl.styles.borders import Border, Side
from movc.province_data import *
from all7.all7_consts import *
from ctt4.ctt4_consts import *
from openpyxl.drawing.image import Image
import re
import pandas as pd
import unittest, logging
import string

class Ctt4:
    def __init__(self, template_file_path, mapping_file_path, year, province):
        self.year = int(year)
        self.template = load_workbook(template_file_path)
        self.thin_border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
        self.province = province
        self.mapping = pd.read_csv(mapping_file_path, encoding = 'ISO-8859-1', sep=';', header = 1)
        self.mapping.columns = ['a', 'b', 'c','d', 'e', 'f', 'g']
        self.img = r"C:\Users\piepalla\PycharmProjects\new_movc\img\istat.png"


    def get_municipalities(self, municipality_code):
        subset = self.mapping[self.mapping.b == int(municipality_code)]
        return list(subset.d)


    def build_ctt4(self, province_name, year, input_dir, output_dir):
        if province_name not in NEW_CODES.keys():
            raise Exception("Wrong province name! Provide one of the five sardinian province name")
        code = NEW_CODES[province_name]
        province_municipalities = self.get_municipalities(code)
        input_dir = os.path.join(input_dir, str(year))
        output_dir = os.path.join(output_dir, str(year))
        allowed_filenames = [p + "_CTT4_" + str(year) + ".xlsx" for p in PROVINCE_FILENAME_MAP[province_name]]
        sect1 = []
        sect2 = []

        try:
            file_names = os.listdir(input_dir)
        except NotADirectoryError:
            print("Wrong input directory!")


        for file in file_names:
            if file not in allowed_filenames:
               continue
            print("current file: " + file)
            file_path = os.path.join(input_dir, file)
            old_ctt4 = load_workbook(file_path)
            sect1 += self.get_matching_rows(province_municipalities, file_path)[CTT4_SHEET1]
            print("partial number of rows: " + str(len(sect1)))
            sect2 += self.get_matching_rows(province_municipalities, file_path)[CTT4_SHEET2]

                # for municipality in province_municipalities:
                #         current_sheet = old_ctt4

        ####### Write to output file ####
        print("Number of rows:" + str(len(sect1)))
        self.writeCTT4(province_name, year, output_dir, sect1, sect2)



    def get_municipality_code(self, municipality_name):
        municipality_code = self.mapping[self.mapping.d == municipality_name].c
        municipality_code = municipality_code.iloc[0]
        municipality_code = str(municipality_code)[-3:]
        return(municipality_code)


    def get_matching_rows(self, province_municipalities, ctt4_file_path):
        old_ctt4 = load_workbook(ctt4_file_path)
        sheet_names = old_ctt4.get_sheet_names()[0:2]
        sect1 = []
        sect2 = []


        for sn in sheet_names:
            current_sheet = old_ctt4.get_sheet_by_name(sn)
            for municipality in province_municipalities:
                for c in current_sheet[MUNICIPALITY_COL_IDX][(MUNICIPALITY_ROW_LOWER_IDX - 1):]:
                    if c.value == "TOTALE":
                        break
                    if municipality == c.value:
                        row = current_sheet[c.row]
                        if (sn == CTT4_SHEET1):
                            sect1.append(row)
                        else:
                            sect2.append(row)

        result = {CTT4_SHEET1: sect1, CTT4_SHEET2: sect2}
        return(result)

    def writeCTT4(self, province_name, year, output_dir, sect1, sect2):
        if province_name not in NEW_CODES.keys():
            raise NameError("Wrong province name. Please, provide one of the five Sardinian Provice name")
        #output_dir = output_dir + str(year) + r"\\"
        output_file_name = PROVINCIAL_SYMBOLS[province_name] + "_ctt4_" + str(year) + ".xlsx"
        output_file = os.path.join(output_dir, output_file_name)

        output_wb = self.template
        sheet_names = output_wb.get_sheet_names()[0:2]
        for sn in sheet_names:
            ws = output_wb.get_sheet_by_name(sn)
            # ### insert logo ###
            img = Image(self.img)
            ws.add_image(img, 'A1')

            #### add province to each sheet
            ws[PROVINCE_CELL_COORDINATE] = province_name.upper() if province_name != "cagliari" else NEW_PROVINCES[0].upper()
            ws[PROVINCE_CODE_COORDINATE] = NEW_CODES[province_name]
            ##########################################################
            ### add year selected
            ws[SELECTED_YEAR_CELL] = "ANNO " + str(year) + " - PER COMUNE"

            #ws["A4"].border = self.thin_border

            row_idx = MUNICIPALITY_ROW_LOWER_IDX
            sect = []
            if sn == CTT4_SHEET1:
                sect = sect1
            else:
                sect = sect2
            for row in sect:
                province_field = ""
                municipality_code = 0
                for c in row:
                    cell_coordinate = c.column + str(row_idx)
                    if c.column == "A":
                        if province_name == "cagliari":
                            province_field = "Città Metropolitana di Cagliari"
                        else:
                            province_field = province_name.capitalize()
                        ws[cell_coordinate] = province_field
                        continue
                    elif c.column == "B":
                        municipality_name = c.value
                        municipality_code = self.get_municipality_code(municipality_name)
                        ws[cell_coordinate] = municipality_name
                        continue
                    elif c.column == "C":
                        ws[cell_coordinate] = municipality_code
                        continue
                    elif (sn == 'Sez 1' and c.column in SHEET1_TOT_CELL) or (sn == "Sez 2" and c.column in SHEET2_TOT_CELL):
                        ws[cell_coordinate] = self.apply_formula(sn, row_idx, c.column)
                        continue
                    ws[cell_coordinate] = c.value
                row_idx += 1

            single_row = sect[0]
            columns = [x.column for x in single_row]
            for col in columns:
                if (sn == CTT4_SHEET1 and col == "BD") or (sn == CTT4_SHEET2 and col == "W"):
                    break
                coordinates = col + str(row_idx)
                if col == "A":
                    ws[coordinates] = province_field
                elif col == "B":
                    ws[coordinates] = "TOTALE"
                elif col == "C":
                    ws[coordinates] = 999
                else:
                    lower_bound = col + str(MUNICIPALITY_ROW_LOWER_IDX)
                    upper_bound = col + str(row_idx - 1)
                    sum_formula = "=SUM(" + lower_bound + ":" + upper_bound + ")"
                    ws[coordinates] = sum_formula
        output_wb.save(output_file)


        return(output_file)


    def add_border(self, sheet_name, worksheet):
        if sheet_name == CTT4_SHEET1:
            cell_list = SHEET1_CELL_LIST
        else:
            cell_list = SHEET2_CELL_LIST

        for coord in cell_list:
            worksheet[coord].border = self.thin_border
        return(worksheet)





    def compute_totals(self, filename, sheetname):
        df = pd.read_excel(filename, sheetname)
        return(df.head(4))




    def apply_formula(self, sheet_name, row_idx, col):
        i = str(row_idx)
        ref_cols =  list(string.ascii_uppercase) + ['AA']
        starting_col = 0
        pace = PACE
        f = ""
        interpolation_var = 0
        if sheet_name == 'Sez 1':
            interpolation_var = 6
            if col == "AB":
                starting_col = ref_cols.index("D")
            elif col == "AC":
                starting_col = ref_cols.index("E")
            elif col == "AD":
                starting_col = ref_cols.index("F")
            elif col == "AE":
                starting_col = ref_cols.index("G")
            else:
                return
            tmp_f = "=SUM(" + ref_cols[starting_col] + "%s," + ref_cols[starting_col + PACE] + "%s," + ref_cols[
                starting_col + PACE * 2] + "%s," + ref_cols[starting_col + PACE * 3] + "%s," + \
                    ref_cols[starting_col + PACE * 4] + "%s," + ref_cols[starting_col + PACE * 5] + "%s)"
        elif sheet_name == 'Sez 2':
            interpolation_var = 3
            if col == "P":
                starting_col = ref_cols.index("D")
            elif col == "Q":
                starting_col = ref_cols.index("E")
            elif col == "R":
                starting_col = ref_cols.index("F")
            elif col == "S" :
                starting_col = ref_cols.index("G")
            else:
                return
            tmp_f = "=SUM(" + ref_cols[starting_col] + "%s," + ref_cols[starting_col + PACE] + "%s," + ref_cols[
                starting_col + PACE * 2] + "%s)"


        f =  tmp_f % tuple([i]*interpolation_var)

        return(f)

